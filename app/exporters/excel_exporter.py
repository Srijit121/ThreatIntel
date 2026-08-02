from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font


class ExcelExporter:
    """Export CVEs to an Excel report."""

    def export(self, cves, filename):
        wb = Workbook()
        ws = wb.active
        ws.title = "CVEs"

        headers = [
            "CVE ID",
            "Severity",
            "CVSS",
            "KEV",
            "KEV Added",
            "KEV Due Date",
            "Vendor",
            "Product",
            "Published",
            "Modified",
            "CWE",
            "Description",
            "Reference URL",
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)

        row = 2

        for cve in cves:
            ws.cell(row=row, column=1).value = cve.id
            ws.cell(row=row, column=2).value = cve.severity
            ws.cell(row=row, column=3).value = cve.cvss_score

            ws.cell(row=row, column=4).value = "YES" if cve.kev else "NO"
            ws.cell(row=row, column=5).value = cve.kev_date
            ws.cell(row=row, column=6).value = cve.kev_due_date

            ws.cell(row=row, column=7).value = cve.vendor
            ws.cell(row=row, column=8).value = cve.product

            ws.cell(row=row, column=9).value = cve.published
            ws.cell(row=row, column=10).value = cve.modified

            ws.cell(row=row, column=11).value = cve.cwe
            ws.cell(row=row, column=12).value = cve.description
            ws.cell(row=row, column=13).value = cve.reference_urls

            row += 1

        Path(filename).parent.mkdir(parents=True, exist_ok=True)

        wb.save(filename)

        # Return the report path so it can be emailed
        return filename
